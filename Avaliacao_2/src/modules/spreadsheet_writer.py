"""Leitura e escrita da Planilha Mestra (.xlsx) com openpyxl.

RN07: "Somente os dados de cadastros APROVADOS devem ser registrados na
Planilha Mestra."
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .exceptions import PlanilhaMestraError

CABECALHO = ["Nome", "CPF", "Data de Nascimento", "Endereço", "Data/Hora do Cadastro"]


def inicializar_planilha_mestra(caminho: Path) -> None:
    """Cria a Planilha Mestra com o cabeçalho padrão, caso ainda não exista."""
    if caminho.exists():
        return

    caminho.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Cadastros Aprovados"
    planilha.append(CABECALHO)
    workbook.save(caminho)


def registrar_cadastro_aprovado(
    dados_cliente: dict, caminho: Path, logger: logging.Logger | None = None
) -> None:
    """Adiciona uma linha à Planilha Mestra com os dados de um cadastro aprovado.

    Levanta PlanilhaMestraError em caso de falha de leitura/gravação
    (cenário de exceção "Erro durante a gravação na Planilha Mestra"), para
    que o chamador NÃO marque o cadastro como concluído e o reprocesse na
    próxima execução, conforme previsto no Mini PDD.
    """
    logger = logger or logging.getLogger("portal_fake_automacao")
    inicializar_planilha_mestra(caminho)

    try:
        workbook = load_workbook(caminho)
        planilha = workbook.active
        planilha.append(
            [
                dados_cliente["nome"],
                dados_cliente["cpf"],
                dados_cliente["data_nascimento"],
                dados_cliente["endereco"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )
        workbook.save(caminho)
    except (KeyError, OSError, PermissionError) as erro:
        raise PlanilhaMestraError(
            f"Falha ao registrar cadastro na Planilha Mestra: {erro}"
        ) from erro

    logger.info("Cadastro de CPF %s registrado na Planilha Mestra", dados_cliente.get("cpf"))
