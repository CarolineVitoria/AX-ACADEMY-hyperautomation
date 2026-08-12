import pytest
from openpyxl import load_workbook

from modules import spreadsheet_writer
from modules.exceptions import PlanilhaMestraError


def test_inicializar_planilha_mestra_cria_arquivo_com_cabecalho(tmp_path):
    caminho = tmp_path / "Planilha_Mestra.xlsx"

    spreadsheet_writer.inicializar_planilha_mestra(caminho)

    assert caminho.exists()
    workbook = load_workbook(caminho)
    planilha = workbook.active
    assert [celula.value for celula in planilha[1]] == spreadsheet_writer.CABECALHO


def test_inicializar_planilha_mestra_nao_sobrescreve_arquivo_existente(tmp_path):
    caminho = tmp_path / "Planilha_Mestra.xlsx"
    spreadsheet_writer.inicializar_planilha_mestra(caminho)
    dados_cliente = {
        "nome": "Maria da Silva",
        "cpf": "52998224725",
        "data_nascimento": "10/05/1990",
        "endereco": "Rua das Flores, 123",
    }
    spreadsheet_writer.registrar_cadastro_aprovado(dados_cliente, caminho)

    spreadsheet_writer.inicializar_planilha_mestra(caminho)

    workbook = load_workbook(caminho)
    assert workbook.active.max_row == 2


def test_registrar_cadastro_aprovado_adiciona_linha_com_os_dados_do_cliente(tmp_path):
    caminho = tmp_path / "Planilha_Mestra.xlsx"
    dados_cliente = {
        "nome": "Maria da Silva",
        "cpf": "52998224725",
        "data_nascimento": "10/05/1990",
        "endereco": "Rua das Flores, 123",
    }

    spreadsheet_writer.registrar_cadastro_aprovado(dados_cliente, caminho)

    workbook = load_workbook(caminho)
    linha = [celula.value for celula in workbook.active[2]]
    assert linha[0:4] == ["Maria da Silva", "52998224725", "10/05/1990", "Rua das Flores, 123"]


def test_registrar_cadastro_com_dados_incompletos_levanta_erro_de_planilha(tmp_path):
    caminho = tmp_path / "Planilha_Mestra.xlsx"
    dados_incompletos = {"nome": "Maria da Silva"}  # faltam cpf, data_nascimento, endereco

    with pytest.raises(PlanilhaMestraError):
        spreadsheet_writer.registrar_cadastro_aprovado(dados_incompletos, caminho)
