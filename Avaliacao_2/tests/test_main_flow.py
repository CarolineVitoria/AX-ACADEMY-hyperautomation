"""Testes de integração do orquestrador (main.py).

O acesso real à caixa de e-mail (Playwright) é substituído por dublês de
teste (mocks), já que depende de um navegador e credenciais reais. As
demais etapas (validação, organização de arquivos e Planilha Mestra)
executam de ponta a ponta, com os caminhos redirecionados para `tmp_path`.
"""

from contextlib import contextmanager

from openpyxl import load_workbook

import main as main_module
from modules.config import Config
from modules.exceptions import CaixaEmailIndisponivelError
from modules.models import STATUS_APROVADO, STATUS_PENDENTE, ResultadoValidacao, Solicitacao


def _redirecionar_caminhos_para_tmp(tmp_path, monkeypatch):
    monkeypatch.setattr(Config, "DOCUMENTOS_OK_DIR", tmp_path / "Documentos_OK")
    monkeypatch.setattr(Config, "DOCUMENTOS_PENDENTES_DIR", tmp_path / "Documentos_Pendentes")
    monkeypatch.setattr(Config, "PLANILHA_MESTRA_PATH", tmp_path / "Planilha_Mestra.xlsx")
    monkeypatch.setattr(Config, "DOWNLOADS_TEMP_DIR", tmp_path / "downloads_temp")
    monkeypatch.setattr(Config, "LOGS_DIR", tmp_path / "logs")


def test_executar_automacao_processa_aprovados_e_pendentes_de_forma_isolada(tmp_path, monkeypatch):
    _redirecionar_caminhos_para_tmp(tmp_path, monkeypatch)

    anexo_aprovado = tmp_path / "Ficha_Cadastro_aprovado.pdf"
    anexo_aprovado.write_text("dummy")
    anexo_pendente = tmp_path / "Ficha_Cadastro_pendente.pdf"
    anexo_pendente.write_text("dummy")

    solicitacao_aprovada = Solicitacao(
        remetente="aprovado@teste.com",
        assunto="Cadastro Portal Fake - 111",
        cpf_assunto="111",
        id_mensagem="1",
        caminhos_anexos={"ficha_cadastro": anexo_aprovado},
    )
    solicitacao_pendente = Solicitacao(
        remetente="pendente@teste.com",
        assunto="Cadastro Portal Fake - 222",
        cpf_assunto="222",
        id_mensagem="2",
        caminhos_anexos={"ficha_cadastro": anexo_pendente},
    )

    @contextmanager
    def sessao_falsa(logger=None):
        yield object()

    def buscar_falso(page, logger=None):
        return [solicitacao_aprovada, solicitacao_pendente]

    def validar_falso(solicitacao, logger=None):
        if solicitacao.cpf_assunto == "111":
            return ResultadoValidacao(
                status=STATUS_APROVADO,
                dados_cliente={
                    "nome": "Fulano de Tal",
                    "cpf": "111",
                    "data_nascimento": "01/01/2000",
                    "endereco": "Rua X, 1",
                },
            )
        return ResultadoValidacao(
            status=STATUS_PENDENTE, pendencias=["Anexo ausente: Documento de Identificação"]
        )

    emails_enviados = []

    def enviar_falso(page, solicitacao, resultado, logger=None):
        emails_enviados.append((solicitacao.cpf_assunto, resultado.status))

    monkeypatch.setattr(main_module, "sessao_webmail", sessao_falsa)
    monkeypatch.setattr(main_module, "buscar_novas_solicitacoes", buscar_falso)
    monkeypatch.setattr(main_module, "validar_solicitacao", validar_falso)
    monkeypatch.setattr(main_module, "enviar_resposta_cliente", enviar_falso)

    codigo_saida = main_module.executar_automacao()

    assert codigo_saida == 0
    assert emails_enviados == [("111", STATUS_APROVADO), ("222", STATUS_PENDENTE)]
    assert (Config.DOCUMENTOS_OK_DIR / "111" / "Ficha_Cadastro_aprovado.pdf").exists()
    assert (Config.DOCUMENTOS_PENDENTES_DIR / "222" / "Ficha_Cadastro_pendente.pdf").exists()

    workbook = load_workbook(Config.PLANILHA_MESTRA_PATH)
    linha_aprovada = [celula.value for celula in workbook.active[2]]
    assert linha_aprovada[0] == "Fulano de Tal"
    assert workbook.active.max_row == 2  # apenas o aprovado foi registrado (RN07)


def test_executar_automacao_com_caixa_indisponivel_encerra_de_forma_controlada(tmp_path, monkeypatch):
    _redirecionar_caminhos_para_tmp(tmp_path, monkeypatch)

    @contextmanager
    def sessao_com_erro(logger=None):
        raise CaixaEmailIndisponivelError("timeout ao conectar na caixa de e-mail")
        yield  # pragma: no cover - necessário para manter a função como gerador

    monkeypatch.setattr(main_module, "sessao_webmail", sessao_com_erro)

    codigo_saida = main_module.executar_automacao()

    # Erro identificado, mensagem registrada em log e processo finalizado de
    # forma controlada (código de saída != 0), sem lançar exceção não tratada.
    assert codigo_saida == 1


def test_executar_automacao_isola_erro_inesperado_de_uma_solicitacao(tmp_path, monkeypatch):
    _redirecionar_caminhos_para_tmp(tmp_path, monkeypatch)

    solicitacao_com_erro = Solicitacao(
        remetente="erro@teste.com",
        assunto="Cadastro Portal Fake - 333",
        cpf_assunto="333",
        id_mensagem="3",
        caminhos_anexos={},
    )

    @contextmanager
    def sessao_falsa(logger=None):
        yield object()

    def buscar_falso(page, logger=None):
        return [solicitacao_com_erro]

    def validar_com_erro_inesperado(solicitacao, logger=None):
        raise RuntimeError("falha inesperada simulada")

    monkeypatch.setattr(main_module, "sessao_webmail", sessao_falsa)
    monkeypatch.setattr(main_module, "buscar_novas_solicitacoes", buscar_falso)
    monkeypatch.setattr(main_module, "validar_solicitacao", validar_com_erro_inesperado)

    # A automação deve finalizar de forma controlada mesmo diante de um erro
    # não previsto durante o processamento de uma solicitação específica.
    codigo_saida = main_module.executar_automacao()

    assert codigo_saida == 0
